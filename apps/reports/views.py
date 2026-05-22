from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.reports.models import Report
from apps.tasks.models import Task
from apps.reports.services import create_weekly_report


@login_required
def index(request):
    reports = Report.objects.select_related("department", "employee", "created_by")
    return render(request, "reports/index.html", {"reports": reports, "title": "Отчёты"})


@login_required
def create(request):
    report = create_weekly_report(request.user)
    messages.success(request, "Отчёт сформирован")
    return redirect("reports:index")


@login_required
def export_excel(request, pk):
    workbook = Workbook()
    sheet = workbook.active
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003C97", end_color="003C97", fill_type="solid")
    bold_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    report = get_object_or_404(Report, pk=pk)
    sheet.title = f"Отчёт {report.pk}"
    
    sheet.append(["Параметр", "Значение"])
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    sheet.append(["Название", report.title])
    sheet.append(["Период", report.get_period_type_display()])
    sheet.append(["Начало", report.period_start.strftime("%d.%m.%Y") if report.period_start else ""])
    sheet.append(["Конец", report.period_end.strftime("%d.%m.%Y") if report.period_end else ""])
    sheet.append(["Отдел", str(report.department) if report.department else "Все отделы"])
    sheet.append(["Создано", report.created_at.strftime("%d.%m.%Y %H:%M")])
    
    sheet.append([""])
    sheet.append(["Содержание (Вывод ИИ):"])
    sheet.cell(row=sheet.max_row, column=1).font = bold_font
    
    content_row = sheet.max_row + 1
    sheet.append([report.content])
    sheet.merge_cells(start_row=content_row, start_column=1, end_row=content_row, end_column=7)
    sheet.cell(row=content_row, column=1).alignment = wrap_alignment
    sheet.row_dimensions[content_row].height = 60
    
    sheet.append([""])
    sheet.append(["Задачи за этот период:"])
    sheet.cell(row=sheet.max_row, column=1).font = bold_font
    
    task_header_row = sheet.max_row + 1
    task_headers = ["ID", "Название", "Статус", "Прогресс", "Приоритет", "Отдел", "Ответственный", "Дедлайн"]
    sheet.append(task_headers)
    for col_num, cell in enumerate(sheet[task_header_row], 1):
        cell.font = header_font
        cell.fill = header_fill

    tasks = Task.objects.filter(
        deadline__gte=report.period_start,
        deadline__lte=report.period_end
    )
    if report.department:
        tasks = tasks.filter(department=report.department)
        
    for t in tasks:
        sheet.append([
            t.id,
            t.title,
            t.get_status_display(),
            f"{t.progress}%" if t.progress is not None else "",
            t.get_priority_display(),
            t.department.name if t.department else "",
            t.responsible.get_full_name() or t.responsible.username if t.responsible else "",
            t.deadline.strftime("%d.%m.%Y") if t.deadline else ""
        ])

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 25
    sheet.column_dimensions['H'].width = 15

    filename = f"report-{report.pk}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

# Create your views here.
